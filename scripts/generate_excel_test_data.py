"""
Generate Excel test data files for parameterized testing
This script creates comprehensive test data workbooks for the Calculator project
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

# Create scripts directory if it doesn't exist
os.makedirs('scripts', exist_ok=True)

def create_addition_sheet():
    """Create Addition test data"""
    data = {
        'TestID': ['ADD-001', 'ADD-002', 'ADD-003', 'ADD-004', 'ADD-005', 'ADD-006', 'ADD-007', 'ADD-008'],
        'a': [5, -5, 5, 0, 100, 0, -1, 1.5],
        'b': [3, -3, -3, 5, 200, 0, 1, 2.5],
        'expected': [8, -8, 2, 5, 300, 0, 0, 4.0],
        'category': ['functional', 'functional', 'functional', 'functional', 'functional', 'functional', 'functional', 'functional'],
        'priority': ['HIGH', 'HIGH', 'HIGH', 'HIGH', 'MEDIUM', 'LOW', 'MEDIUM', 'MEDIUM'],
        'description': ['Positive numbers', 'Negative numbers', 'Mixed numbers', 'With zero', 'Large numbers', 'Zero + zero', 'Negative to positive', 'Float addition'],
        'author': ['Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham'],
        'date': [datetime.now().date()] * 8,
    }
    return pd.DataFrame(data)

def create_subtraction_sheet():
    """Create Subtraction test data"""
    data = {
        'TestID': ['SUB-001', 'SUB-002', 'SUB-003', 'SUB-004', 'SUB-005', 'SUB-006', 'SUB-007'],
        'a': [10, -5, 5, 5, 0, 100, 3.5],
        'b': [3, -3, -3, 0, 5, 100, 1.5],
        'expected': [7, -2, 8, 5, -5, 0, 2.0],
        'category': ['functional', 'functional', 'functional', 'functional', 'functional', 'functional', 'functional'],
        'priority': ['HIGH', 'HIGH', 'HIGH', 'MEDIUM', 'MEDIUM', 'LOW', 'MEDIUM'],
        'description': ['Positive numbers', 'Negative numbers', 'Mixed numbers', 'With zero', 'Zero minus positive', 'Same numbers', 'Float subtraction'],
        'author': ['Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham'],
        'date': [datetime.now().date()] * 7,
    }
    return pd.DataFrame(data)

def create_multiplication_sheet():
    """Create Multiplication test data"""
    data = {
        'TestID': ['MUL-001', 'MUL-002', 'MUL-003', 'MUL-004', 'MUL-005', 'MUL-006', 'MUL-007', 'MUL-008'],
        'a': [5, -5, 5, 5, 0, 10, 2.5, -1],
        'b': [3, -3, -3, 0, 0, 10, 4, -1],
        'expected': [15, 15, -15, 0, 0, 100, 10.0, 1],
        'category': ['functional', 'functional', 'functional', 'functional', 'boundary', 'functional', 'functional', 'functional'],
        'priority': ['HIGH', 'HIGH', 'HIGH', 'HIGH', 'HIGH', 'MEDIUM', 'MEDIUM', 'MEDIUM'],
        'description': ['Positive numbers', 'Negative numbers', 'Mixed numbers', 'Multiply by zero', 'Zero times zero', 'Large numbers', 'Float multiplication', 'Negative × negative'],
        'author': ['Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham'],
        'date': [datetime.now().date()] * 8,
    }
    return pd.DataFrame(data)

def create_division_sheet():
    """Create Division test data"""
    data = {
        'TestID': ['DIV-001', 'DIV-002', 'DIV-003', 'DIV-004', 'DIV-005', 'DIV-006', 'DIV-007', 'DIV-008', 'DIV-009'],
        'a': [10, -10, 10, 5, 6, 100, 2.5, 0, 7],
        'b': [2, -2, -2, 2, 3, 10, 2, 5, 0],
        'expected': [5.0, 5.0, -5.0, 2.5, 2.0, 10.0, 1.25, 0.0, 'ERROR'],
        'category': ['functional', 'functional', 'functional', 'functional', 'functional', 'functional', 'functional', 'functional', 'boundary'],
        'priority': ['HIGH', 'HIGH', 'HIGH', 'HIGH', 'HIGH', 'MEDIUM', 'MEDIUM', 'LOW', 'CRITICAL'],
        'description': ['Positive numbers', 'Negative numbers', 'Mixed numbers', 'Result as float', 'Whole number result', 'Large numbers', 'Float division', 'Zero divided', 'Division by zero'],
        'author': ['Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham'],
        'date': [datetime.now().date()] * 9,
    }
    return pd.DataFrame(data)

def create_security_sheet():
    """Create Security test data"""
    data = {
        'TestID': ['SEC-001', 'SEC-002', 'SEC-003', 'SEC-004', 'SEC-005', 'SEC-006', 'SEC-007'],
        'input': [
            "5' OR '1'='1",
            "__import__('os').system('rm -rf /')",
            "<script>alert('XSS')</script>",
            "; DROP TABLE users;",
            "你好123",
            "5injection_nullbyte",
            "9" * 1000,
        ],
        'operation': ['add', 'subtract', 'multiply', 'divide', 'power', 'modulo', 'square_root'],
        'should_fail': [True, True, True, True, True, True, True],
        'category': ['security', 'security', 'security', 'security', 'security', 'security', 'security'],
        'priority': ['CRITICAL', 'CRITICAL', 'CRITICAL', 'CRITICAL', 'HIGH', 'HIGH', 'MEDIUM'],
        'threat_type': ['SQL Injection', 'Code Injection', 'XSS', 'Command Injection', 'Encoding Attack', 'Null Byte Injection', 'Buffer Overflow'],
        'author': ['Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham'],
        'date': [datetime.now().date()] * 7,
    }
    return pd.DataFrame(data)

def create_boundary_sheet():
    """Create Boundary test data"""
    data = {
        'TestID': ['BND-001', 'BND-002', 'BND-003', 'BND-004', 'BND-005', 'BND-006', 'BND-007', 'BND-008'],
        'operation': ['divide', 'modulo', 'add', 'add', 'square_root', 'square_root', 'divide', 'multiply'],
        'a': [5, 10, 1e308, -1e308, -1, -4, 1e-100, 1e100],
        'b': [0, 0, 0, 0, None, None, 0, 1e100],
        'exception': ['ValueError', 'ValueError', 'OK', 'OK', 'ValueError', 'ValueError', 'ValueError', 'OK'],
        'category': ['boundary', 'boundary', 'boundary', 'boundary', 'boundary', 'boundary', 'boundary', 'boundary'],
        'priority': ['CRITICAL', 'CRITICAL', 'HIGH', 'HIGH', 'HIGH', 'HIGH', 'CRITICAL', 'MEDIUM'],
        'description': ['Division by zero', 'Modulo by zero', 'Extreme positive', 'Extreme negative', 'Negative square root', 'Negative square root', 'Extreme divide by zero', 'Very large numbers'],
        'author': ['Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham', 'Shubham'],
        'date': [datetime.now().date()] * 8,
    }
    return pd.DataFrame(data)

def create_performance_sheet():
    """Create Performance test data"""
    data = {
        'TestID': ['PERF-001', 'PERF-002', 'PERF-003', 'PERF-004'],
        'operation': ['add', 'multiply', 'square_root', 'power'],
        'a': [1e100, 999, 10000, 2],
        'b': [1e100, 999, None, 100],
        'max_time_seconds': [1.0, 1.0, 1.0, 1.0],
        'category': ['performance', 'performance', 'performance', 'performance'],
        'priority': ['HIGH', 'HIGH', 'HIGH', 'HIGH'],
        'description': ['Large number addition', 'Large number multiplication', 'Large square root', 'Power calculation'],
        'author': ['Shubham', 'Shubham', 'Shubham', 'Shubham'],
        'date': [datetime.now().date()] * 4,
    }
    return pd.DataFrame(data)

def apply_formatting(writer, sheetname, df):
    """Apply Excel formatting to make it look professional"""
    workbook = writer.book
    worksheet = writer.sheets[sheetname]
    
    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, value in enumerate(df.columns.values, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = value
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data formatting
    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    """Generate all Excel test data files"""
    output_file = 'test_data/calculator_tests.xlsx'
    
    print(f"Generating Excel test data: {output_file}\n")
    
    # Create data sheets
    sheets = {
        'Addition': create_addition_sheet(),
        'Subtraction': create_subtraction_sheet(),
        'Multiplication': create_multiplication_sheet(),
        'Division': create_division_sheet(),
        'Security': create_security_sheet(),
        'Boundary': create_boundary_sheet(),
        'Performance': create_performance_sheet(),
    }
    
    # Write to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            apply_formatting(writer, sheet_name, df)
            print(f"✓ Sheet '{sheet_name}': {len(df)} test cases")
    
    print(f"\n✅ Excel file created: {output_file}")
    print(f"📊 Total test cases: {sum(len(df) for df in sheets.values())}")
    print(f"\nSheet summary:")
    for sheet_name, df in sheets.items():
        print(f"   - {sheet_name}: {len(df)} tests")

if __name__ == "__main__":
    main()
